from openpyxl import Workbook
import json

write_wb = Workbook()
write_ws = write_wb.create_sheet("sample")
write_ws = write_wb.active

with open("crawler\yes24_2022_0221_1218_16.json", encoding="utf-8") as json_file:
    json_data = json.load(json_file)
    row = 1
    for data in json_data:
        print(json_data[data]["title"])
        write_ws.cell(row=row, column=1, value=json_data[data]["title"])
        write_ws.cell(row=row, column=2, value=json_data[data]["url"])
        write_ws.cell(row=row, column=3, value=json_data[data]["rank"])
        write_ws.cell(row=row, column=4, value=str(json_data[data]["publisher"]))
        write_ws.cell(row=row, column=5, value=json_data[data]["publish_date"])
        write_ws.cell(row=row, column=6, value=json_data[data]["right_price"])
        write_ws.cell(row=row, column=7, value=json_data[data]["isbn"])
        write_ws.cell(row=row, column=8, value=json_data[data]["page"])
        write_ws.cell(row=row, column=9, value=json_data[data]["sales_point"])
        write_ws.cell(row=row, column=10, value=str(json_data[data]["author"]))
        write_ws.cell(row=row, column=11, value=str(json_data[data]["tags"]))
        row += 1


write_wb.save("./test.xlsx")
